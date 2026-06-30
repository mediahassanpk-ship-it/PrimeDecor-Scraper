"""
Exporter module for PrimeDecor Scraper
Handles exporting scraped data to Excel and other formats
"""

import os
from typing import List, Dict, Any, Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import EXCEL_COLUMNS, OUTPUT_DIR, OUTPUT_PATH
from utils import (
    get_logger, strip_html, clean_text, normalize_url,
    format_price, format_bool, format_date, parse_tags,
    parse_collections, format_inventory_quantity, format_available,
    extract_image_urls, format_variant_name, format_weight_unit
)


class DataExporter:
    """Export scraped product data to Excel"""
    
    def __init__(self, output_path: str = OUTPUT_PATH):
        """
        Initialize exporter
        
        Args:
            output_path: Path to output Excel file
        """
        self.output_path = output_path
        self.logger = get_logger("exporter")
        self.products_data = []
        
    def prepare_product_row(self, product: Dict[str, Any], variant: Dict[str, Any]) -> Dict[str, Any]:
        """
        Prepare product-variant row for export
        
        Args:
            product: Product data
            variant: Variant data
        
        Returns:
            Formatted row dictionary
        """
        try:
            product_url = f"https://primedecor.pk/products/{product.get('handle', '')}"
            
            # Extract images
            featured_image = ""
            all_images = []
            
            if product.get('featured_image'):
                featured_image = product['featured_image'].get('src', '')
            
            if product.get('images'):
                all_images = extract_image_urls(product['images'])
            
            # Prepare variant-specific data
            sku = variant.get('sku', '')
            barcode = variant.get('barcode', '')
            price = format_price(variant.get('price'))
            compare_at_price = format_price(variant.get('compare_at_price'))
            available = format_available(variant)
            inventory_qty = variant.get('inventory_quantity', 0)
            weight = variant.get('weight')
            weight_unit = format_weight_unit(variant.get('weight_unit', 'kg'))
            taxable = format_bool(variant.get('taxable', False))
            requires_shipping = format_bool(variant.get('requires_shipping', True))
            
            # Extract options
            option1 = variant.get('option1', '')
            option2 = variant.get('option2', '')
            option3 = variant.get('option3', '')
            
            # Get variant name
            variant_name = format_variant_name(variant, product.get('title', ''))
            
            # Extract SEO data
            seo_title = ""
            seo_description = ""
            if product.get('seo'):
                seo_title = product['seo'].get('title', '')
                seo_description = product['seo'].get('description', '')
            
            # Prepare row
            row = {
                "Product ID": str(product.get('id', '')),
                "Product Handle": clean_text(product.get('handle', '')),
                "Product URL": product_url,
                "Product Name": clean_text(product.get('title', '')),
                "Vendor": clean_text(product.get('vendor', '')),
                "Category (Product Type)": clean_text(product.get('product_type', '')),
                "Collections": parse_collections(product.get('collections', [])),
                "Tags": ", ".join(parse_tags(product.get('tags', ''))),
                "Description (Plain Text)": strip_html(product.get('body_html', '')),
                "Published": format_date(product.get('published_at', '')),
                "Status": clean_text(product.get('status', '')),
                "Variant ID": str(variant.get('id', '')),
                "Variant Name": variant_name,
                "Option1": clean_text(option1),
                "Option2": clean_text(option2),
                "Option3": clean_text(option3),
                "SKU": clean_text(sku),
                "Barcode": clean_text(barcode),
                "Price": price,
                "Compare At Price": compare_at_price,
                "Available": available,
                "Inventory Quantity": inventory_qty,
                "Weight": weight,
                "Weight Unit": weight_unit,
                "Taxable": taxable,
                "Requires Shipping": requires_shipping,
                "Featured Image": featured_image,
                "All Image URLs": " | ".join(all_images) if all_images else "",
                "SEO Title": clean_text(seo_title),
                "SEO Description": clean_text(seo_description),
            }
            
            return row
            
        except Exception as e:
            self.logger.error(f"Error preparing product row: {e}")
            return {}
    
    def add_product(self, product: Dict[str, Any]):
        """
        Add product to export data
        
        Args:
            product: Product data with variants
        """
        try:
            variants = product.get('variants', [])
            
            if not variants:
                # Product with no variants
                variant_row = self.prepare_product_row(product, {})
                if variant_row:
                    self.products_data.append(variant_row)
            else:
                # Create row for each variant
                for variant in variants:
                    variant_row = self.prepare_product_row(product, variant)
                    if variant_row:
                        self.products_data.append(variant_row)
            
        except Exception as e:
            self.logger.error(f"Error adding product: {e}")
    
    def add_products_batch(self, products: List[Dict[str, Any]]):
        """
        Add batch of products to export data
        
        Args:
            products: List of product data
        """
        for product in products:
            self.add_product(product)
    
    def export_to_excel(self, output_path: Optional[str] = None) -> str:
        """
        Export data to Excel file
        
        Args:
            output_path: Path to output file (uses default if not provided)
        
        Returns:
            Path to created file
        """
        try:
            output_path = output_path or self.output_path
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path) or OUTPUT_DIR, exist_ok=True)
            
            if not self.products_data:
                self.logger.warning("No product data to export")
                return ""
            
            # Create DataFrame
            df = pd.DataFrame(self.products_data)
            
            # Reorder columns according to config
            columns_to_use = [col for col in EXCEL_COLUMNS if col in df.columns]
            df = df[columns_to_use]
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Products', index=False)
                
                # Format worksheet
                worksheet = writer.sheets['Products']
                self._format_worksheet(worksheet)
            
            self.logger.info(f"Excel file created: {output_path}")
            self.logger.info(f"Total rows exported: {len(self.products_data)}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def export_to_csv(self, output_path: Optional[str] = None) -> str:
        """
        Export data to CSV file
        
        Args:
            output_path: Path to output file
        
        Returns:
            Path to created file
        """
        try:
            if not output_path:
                output_path = self.output_path.replace('.xlsx', '.csv')
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path) or OUTPUT_DIR, exist_ok=True)
            
            if not self.products_data:
                self.logger.warning("No product data to export")
                return ""
            
            # Create DataFrame
            df = pd.DataFrame(self.products_data)
            
            # Reorder columns
            columns_to_use = [col for col in EXCEL_COLUMNS if col in df.columns]
            df = df[columns_to_use]
            
            # Export to CSV
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            self.logger.info(f"CSV file created: {output_path}")
            self.logger.info(f"Total rows exported: {len(self.products_data)}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error exporting to CSV: {e}")
            raise
    
    def _format_worksheet(self, worksheet):
        """
        Format Excel worksheet with styles
        
        Args:
            worksheet: openpyxl worksheet object
        """
        try:
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Format header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Set column widths and format data rows
            column_widths = {
                'A': 15,  # Product ID
                'B': 20,  # Product Handle
                'C': 25,  # Product URL
                'D': 25,  # Product Name
                'E': 15,  # Vendor
                'F': 20,  # Category
                'G': 25,  # Collections
                'H': 20,  # Tags
                'I': 30,  # Description
                'J': 15,  # Published
                'K': 12,  # Status
                'L': 15,  # Variant ID
                'M': 20,  # Variant Name
                'N': 15,  # Option1
                'O': 15,  # Option2
                'P': 15,  # Option3
                'Q': 12,  # SKU
                'R': 15,  # Barcode
                'S': 12,  # Price
                'T': 15,  # Compare At Price
                'U': 12,  # Available
                'V': 15,  # Inventory Qty
                'W': 12,  # Weight
                'X': 12,  # Weight Unit
                'Y': 10,  # Taxable
                'Z': 15,  # Requires Shipping
                'AA': 25, # Featured Image
                'AB': 30, # All Image URLs
                'AC': 25, # SEO Title
                'AD': 30, # SEO Description
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Format data rows
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = left_alignment
                    
                    # Format specific columns
                    col_letter = cell.column_letter
                    
                    # Center align specific columns
                    if col_letter in ['A', 'L', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Format prices
                    if col_letter in ['S', 'T']:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0.00'
                    
                    # Format URLs as hyperlinks (for featured image)
                    if col_letter == 'AA' and cell.value and str(cell.value).startswith('http'):
                        cell.hyperlink = cell.value
                        cell.font = Font(color="0563C1", underline="single")
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
        except Exception as e:
            self.logger.error(f"Error formatting worksheet: {e}")
    
    def get_stats(self) -> Dict[str, Any]:
        """
        Get export statistics
        
        Returns:
            Dictionary with statistics
        """
        if not self.products_data:
            return {
                'total_rows': 0,
                'total_products': 0,
                'total_variants': 0,
            }
        
        df = pd.DataFrame(self.products_data)
        
        return {
            'total_rows': len(df),
            'total_products': df['Product ID'].nunique() if 'Product ID' in df.columns else 0,
            'total_variants': df['Variant ID'].nunique() if 'Variant ID' in df.columns else 0,
            'avg_price': df['Price'].mean() if 'Price' in df.columns else 0,
            'min_price': df['Price'].min() if 'Price' in df.columns else 0,
            'max_price': df['Price'].max() if 'Price' in df.columns else 0,
        }
    
    def clear(self):
        """Clear all exported data"""
        self.products_data = []
        self.logger.debug("Export data cleared")
